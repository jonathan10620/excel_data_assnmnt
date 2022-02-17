import sre_parse
from nbformat import write
import math
import random
from openpyxl import load_workbook
import csv
from csv import excel, reader
from pprint import pprint
from helpers import clean_age, convert_height_to_inches, hide_age
from random import choice, randrange
from openpyxl.styles import PatternFill
from datetime import datetime
from datetime import timedelta
import datedelta


excel_file = "main.xlsx"


def write_description_cells(sheet_name, col_name):
    try:
        col_dict_csv = open(f"dictionary/{col_name}.csv")
    except:
        col_dict_csv = None
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # read in values
    try:
        csvreader = csv.reader(col_dict_csv)
    except:
        pass

    # create a dictionary of from passed csv
    code_dict = {}
    try:
        for row in csvreader:
            code_dict[row[0]] = row[1]
    except:
        pass

    # if col is admission_type:
    if col_name == "admission_type":
        # read each B cell code and insert the description in C cell
        for n, cell in enumerate(ws["B"]):
            if ws["C" + str(n + 1)].value is None:
                if str(cell.value) in code_dict:
                    ws["C" + str(n + 1)].value = code_dict[str(cell.value)]
                else:
                    ws["C" + str(n + 1)].value = "N/A"
    elif col_name == "admit_src":
        for n, cell in enumerate(ws["D"]):
            if ws["E" + str(n + 1)].value is None:
                if str(cell.value) in code_dict:
                    ws["E" + str(n + 1)].value = code_dict[str(cell.value)]
                else:
                    ws["E" + str(n + 1)].value = "N/A"
    elif col_name == "race":
        for n, cell in enumerate(ws["U"]):
            if str(cell.value) in code_dict:
                print(cell.value)
                print(type(cell.value))
                ws["V" + str(n + 1)].value = code_dict[str(cell.value)]
            else:
                ws["V" + str(n + 1)].value = "N/A"
    elif col_name == "ethnicity":
        for n, cell in enumerate(ws["W"]):
            if str(cell.value) in code_dict:
                ws["X" + str(n + 1)].value = code_dict[str(cell.value)]
            else:
                ws["X" + str(n + 1)].value = "N/A"
    elif col_name == "insurance":
        for n, cell in enumerate(ws["AH"]):
            if str(cell.value) in code_dict:
                ws["AI" + str(n + 1)].value = code_dict[str(cell.value)]
            else:
                ws["AI" + str(n + 1)].value = "N/A"
    elif col_name == "icd10":
        for n, cell in enumerate(ws["AQ"]):
            if str(cell.value) in code_dict:
                ws["AR" + str(n + 1)].value = code_dict[str(cell.value)].strip('"')
                ws["AS" + str(n + 1)].value = cell.value[0]
                ws["AT" + str(n + 1)].value = cell.value[:3]
            else:
                ws["AR" + str(n + 1)].value = "N/A"
                ws["AS" + str(n + 1)].value = "N/A"
                ws["AT" + str(n + 1)].value = "N/A"
                ws["AR" + str(n + 1)].fill = PatternFill(
                    start_color="00FA92", fill_type="solid"
                )
                ws["AS" + str(n + 1)].fill = PatternFill(
                    start_color="00FA92", fill_type="solid"
                )
                ws["AT" + str(n + 1)].fill = PatternFill(
                    start_color="00FA92", fill_type="solid"
                )
    elif col_name == "sev":
        sev_desc = {"1": "Minor", "2": "Moderate", "3": "Major", "4": "Extreme"}
        for n, cell in enumerate(ws["BD"]):
            if str(cell.value) in sev_desc:
                ws["BE" + str(n + 1)].value = sev_desc[str(cell.value)]

    wb.save(excel_file)


def name_col(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for first, last in zip(ws["G"], ws["I"]):
        if "_" in str(first.value):
            continue
        ws["J" + str(first.row)].value = str(first.value) + " " + str(last.value)

    wb.save(excel_file)


def city_county_column(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    with open("dictionary/tx_zips.csv") as read_obj:
        # pass the file object to reader() to get the reader object
        reader = csv.reader(read_obj)

        data = list(reader)

        zip_dict = {}

        for row in data:
            zip_dict[row[0]] = [row[1], row[2]]

    for zip in ws["N"]:
        if zip.value == "City" or zip.value == "County":
            continue
        if str(zip.value) in zip_dict:
            ws["P" + str(zip.row)].value = zip_dict[str(zip.value)][0]
            ws["L" + str(zip.row)].value = zip_dict[str(zip.value)][1]
        else:
            ws["P" + str(zip.row)].value = "N/A"
            ws["L" + str(zip.row)].value = "N/A"
    wb.save(excel_file)


def died(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, cell in enumerate(ws["Q"]):
        if ws["R" + str(n + 1)].value is None:
            if cell.value in [
                48,
                42,
                40,
                20,
            ]:
                ws["R" + str(n + 1)].value = "Y"
            else:
                ws["R" + str(n + 1)].value = "N"

    wb.save(excel_file)


def newborn(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, cell in enumerate(ws["B"]):
        print(cell.value)
        print(type(cell.value))
        if ws["S" + str(n + 1)].value is None:
            if cell.value == 4:
                ws["S" + str(n + 1)].value = "Y"
            else:
                ws["S" + str(n + 1)].value = "N"

    wb.save(excel_file)


def salutation(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, cell in enumerate(ws["T"]):
        if cell.value == "M":
            ws["Z" + str(n + 1)].value = "Mr."
        elif cell.value == "F" and ws["Y" + str(n + 1)].value == "Married":
            ws["Z" + str(n + 1)].value = "Mrs."
        elif cell.value == "F":
            ws["Z" + str(n + 1)].value = "Ms."
        else:
            ws["Z" + str(n + 1)].value = "N/A"

    wb.save(excel_file)


def DC_date(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, cell in enumerate(ws["AA"]):
        length_of_stay = int(ws["AE" + str(n + 1)].value)
        discharge = cell.value + datetime.timedelta(days=length_of_stay)

        ws["AB" + str(n + 1)].value = discharge.strftime("%-m/%-d/%Y")
        ws["AC" + str(n + 1)].value = discharge.strftime("%A")

        import pandas as pd

        business_days = 0
        start_date = cell.value

        datelist = pd.date_range(start_date, periods=length_of_stay + 1).tolist()

        for day in datelist:
            if day.weekday() in [0, 1, 2, 3, 4]:
                business_days += 1

        ws["AD" + str(n + 1)].value = business_days

    wb.save(excel_file)


def age(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    age_dict = {}

    # with open('dictionary/age.csv') as read_obj:
    #     csvreader = csv.reader(read_obj)
    #     for line in csvreader:
    #         age_dict[int(line[0])] = clean_age(line[1])

    # # pat code int, csv code str
    # for n, cell in enumerate(ws['AF']):
    #     if sheet_name == 'main' and n == 0:
    #         continue
    #     if ws['AG' + str(n+1)].value is None:
    #         if int(cell.value) in age_dict:
    #             ws["AG" + str(n+1)].value = age_dict[int(cell.value)]
    #         else:
    #             ws["AG" + str(n+1)].value = str(hide_age(int(cell.value))) + ' years'
    #             ws["AG" + str(n+1)].fill = PatternFill(start_color='00FA92', fill_type='solid')

    for n, (admit, age) in enumerate(zip(ws["AA"], ws["AG"])):
        if age.value == "Age":
            continue
        admit = admit.value

        age = int(("".join([x for x in age.value if x.isdigit()]).strip()))

        if age > 99:
            birthday = admit - datedelta.datedelta(years=0)
        else:
            birthday = admit - datedelta.datedelta(years=age)
        # except Exception as e:
        #     birthday = admit - datedelta.datedelta(years=0)
        
        astro_dict = {
            "Aries": 1,
            "Taurus": 2,
            "Gemini": 3,
            "Cancer": 4,
            "Leo": 5,
            "Virgo": 6,
            "Libra": 7,
            "Scorpio": 8,
            "Sagittarius": 9,
            "Capricorn": 10,
            "Aquarius": 11,
            "Pisces": 12,
        }

        astro_dict2 = {v: k for k, v in astro_dict.items()}

        for cell in ws["BS"]:
            # ws["BS" + str(n + 1)].value = birthday.strftime("%-m/%-d/%Y")
            ws["BT" + str(n + 1)].value = astro_dict2[int(birthday.strftime("%-m"))]
    wb.save(excel_file)


def charge(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, cell in enumerate(ws["AJ"]):
        if cell.value == "TOTAL_CHARGES":
            continue
        else:
            # integer
            ws["AK" + str(n + 1)].value = int(cell.value)
            # round up
            ws["AL" + str(n + 1)].value = math.ceil(cell.value)
            # round down
            ws["AM" + str(n + 1)].value = math.floor(cell.value)

    for n, (los, cost) in enumerate(zip(ws["AE"], ws["AJ"])):
        if los.value == "LENGTH_OF_STAY":
            continue
        else:
            ws["AP" + str(n + 1)].value = float("{:.2f}".format(cost.value / los.value))

    wb.save(excel_file)


def height(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, (height, weight) in enumerate(zip(ws["BJ"], ws["BL"])):
        if height.value == "Height" or weight.value == "Weight":
            continue
        ws["BK" + str(n + 1)].value = convert_height_to_inches(height.value)
        ws["BM" + str(n + 1)].value = convert_height_to_inches(height.value) * 2.54
        # wt kg
        ws["BN" + str(n + 1)].value = float("{:.2f}".format(weight.value * 0.45359237))

        # BMI = weight (lb) / [height (in)]2 x 703
        ws["BO" + str(n + 1)].value = int(
            weight.value / (convert_height_to_inches(height.value) ** 2) * 703
        )

    wb.save(excel_file)


def bmi_desc(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for n, cell in enumerate(ws["BO"]):
        if cell.value == "BMI":
            continue
        elif cell.value < 18.5:
            ws["BP" + str(n + 1)].value = "Underweight"
        elif cell.value >= 18.5 and cell.value <= 24.9:
            ws["BP" + str(n + 1)].value = "Normal"
        elif cell.value >= 25 and cell.value <= 29.9:
            ws["BP" + str(n + 1)].value = "Overweight"
        elif cell.value >= 30:
            ws["BP" + str(n + 1)].value = "Obese"

    wb.save(excel_file)


def vitals(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    # HR
    for n, (bmi, HR) in enumerate(zip(ws["BP"], ws["BQ"])):
        if bmi.value == "BMI":
            continue
        elif str(bmi.value) == "Overweight":
            HR.value = randrange(80, 100)
        elif str(bmi.value) == "Obese":
            HR.value = randrange(100, 120)
        elif str(bmi.value) == "Underweight":
            HR.value = randrange(45, 65)
        elif str(bmi.value) == "Normal":
            HR.value = randrange(60, 80)
    wb.save(excel_file)


def temp(sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    with open("cleaned_fever.txt", "a+") as f:
        f.seek(0)
        icd10_arr = [x.strip() for x in f.read().split("\n") if x != ""]
        final_icd = []
        for dx in icd10_arr:
            l = dx.split(",")
            for i in l:
                if i[0].isalpha():
                    final_icd.append(i.strip())

        # Ffinal_icd represents text of key words incdicating infection (high temp)

        for n, cell in enumerate(ws["AR"]):
            if cell.value == "ICD_10_desc":
                continue

            for word in cell.value.split(" "):
                if str(word) in final_icd:
                    ws["BR" + str(n + 1)].value = round(random.uniform(99.5, 101.2), 2)
                elif "infection" in word:
                    ws["BR" + str(n + 1)].value = round(random.uniform(98.9, 99.4), 2)
                elif "fever" in word:
                    ws["BR" + str(n + 1)].value = round(random.uniform(99.0, 102.1), 2)
                else:
                    ws["BR" + str(n + 1)].value = round(random.uniform(97.2, 98.4), 2)

        wb.save(excel_file)
